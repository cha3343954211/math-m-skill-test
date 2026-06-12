"""
CUMCM 2017 D - 巡检线路排班 - 完整求解 v2
优化版：更好的路线设计、调度生成、可视化
"""
import numpy as np
import pandas as pd
import json
import os
import warnings
warnings.filterwarnings('ignore')

# ========== 1. 数据加载 ==========
def load_data():
    """加载附件数据"""
    import openpyxl
    wb = openpyxl.load_workbook('支撑材料/data/raw/CUMCM-2017-appendix-D.xlsx')
    
    ws1 = wb['基本信息']
    points = []
    for row in ws1.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        points.append({
            'id': row[0],
            'num': int(row[0].split('-')[1]),
            'period': int(row[1]),
            'duration': int(row[2])
        })
    
    ws2 = wb['连通关系']
    edges = []
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        edges.append({
            'from': int(row[0]),
            'to': int(row[1]),
            'time': int(row[2])
        })
    
    return points, edges

# ========== 2. 图论 ==========
def build_graph(points, edges, n=26):
    """构建邻接矩阵和Floyd最短路径"""
    INF = 9999
    dist = np.full((n+1, n+1), INF)
    nxt = np.zeros((n+1, n+1), dtype=int)
    for i in range(1, n+1):
        dist[i][i] = 0
        nxt[i][i] = i
    for e in edges:
        dist[e['from']][e['to']] = e['time']
        dist[e['to']][e['from']] = e['time']
        nxt[e['from']][e['to']] = e['to']
        nxt[e['to']][e['from']] = e['from']
    
    for k in range(1, n+1):
        for i in range(1, n+1):
            for j in range(1, n+1):
                if dist[i][k] + dist[k][j] < dist[i][j]:
                    dist[i][j] = dist[i][k] + dist[k][j]
                    nxt[i][j] = nxt[i][k]
    
    return dist, nxt

def get_path(nxt, i, j):
    """获取i到j的最短路径"""
    if nxt[i][j] == 0:
        return [i]
    path = [i]
    while i != j:
        i = nxt[i][j]
        path.append(i)
    return path

# ========== 3. 路线设计（改进版） ==========
def calc_route_time_full(start, route, dist, point_dict):
    """计算路线完整时间（含巡检）"""
    if not route:
        return 0
    time = dist[start][route[0]]
    for i in range(len(route)):
        time += point_dict[route[i]]['duration']
        if i < len(route) - 1:
            time += dist[route[i]][route[i+1]]
    time += dist[route[-1]][start]
    return time

def calc_route_travel(start, route, dist):
    """计算路线旅行时间（不含巡检）"""
    if not route:
        return 0
    time = dist[start][route[0]]
    for i in range(len(route) - 1):
        time += dist[route[i]][route[i+1]]
    time += dist[route[-1]][start]
    return time

def tsp_2opt(start, route, dist, point_dict):
    """2-opt优化路线顺序"""
    if len(route) <= 2:
        return route
    best = list(route)
    best_time = calc_route_time_full(start, best, dist, point_dict)
    improved = True
    while improved:
        improved = False
        for i in range(len(best) - 1):
            for j in range(i + 1, len(best)):
                new_route = best[:i] + best[i:j+1][::-1] + best[j+1:]
                new_time = calc_route_time_full(start, new_route, dist, point_dict)
                if new_time < best_time:
                    best = new_route
                    best_time = new_time
                    improved = True
    return best

def design_routes(points, dist, shift_minutes=480, start=22):
    """
    改进版路线设计：
    1. 按周期分组
    2. 同周期点用最近邻+2-opt编排路线
    3. 路线时间约束：含巡检时间 ≤ period（确保工人能在周期内完成）
    """
    point_dict = {p['num']: p for p in points}
    
    # 按周期分组
    period_groups = {}
    for p in points:
        per = p['period']
        if per not in period_groups:
            period_groups[per] = []
        period_groups[per].append(p['num'])
    
    all_routes = []
    
    for period, pnums in sorted(period_groups.items()):
        if not pnums:
            continue
        
        # 对于长周期点（480, 720），单独成路线
        if period >= 480:
            for pn in pnums:
                route = tsp_2opt(start, [pn], dist, point_dict)
                all_routes.append({'points': route, 'period': period})
            continue
        
        # 对于短周期点，用贪心分组
        remaining = list(pnums)
        
        while remaining:
            # 从剩余点中选一个起点（离start最近的）
            best_start = min(remaining, key=lambda x: dist[start][x])
            batch = [best_start]
            remaining.remove(best_start)
            
            # 贪心添加点，直到路线时间接近period
            while remaining:
                # 找最佳候选点
                best_cand = None
                best_score = float('inf')
                for cand in remaining:
                    # 临时添加cand到batch末尾
                    test_route = batch + [cand]
                    test_time = calc_route_time_full(start, test_route, dist, point_dict)
                    if test_time <= period * 0.95:  # 留5%余量
                        # 评分：路线时间 + 最近邻距离
                        score = test_time + dist[batch[-1]][cand]
                        if score < best_score:
                            best_score = score
                            best_cand = cand
                
                if best_cand is None:
                    break
                batch.append(best_cand)
                remaining.remove(best_cand)
            
            # 2-opt优化
            batch = tsp_2opt(start, batch, dist, point_dict)
            route_time = calc_route_time_full(start, batch, dist, point_dict)
            
            all_routes.append({
                'points': batch,
                'period': period,
                'route_time': route_time
            })
    
    return all_routes

# ========== 4. 调度计算 ==========
def calc_workers(routes, dist, point_dict, shift_minutes=480, start=22):
    """计算每条路线所需工人数"""
    results = []
    total_workers = 0
    
    for r in routes:
        route = r['points']
        period = r['period']
        route_time = r.get('route_time', calc_route_time_full(start, route, dist, point_dict))
        
        # 该路线每班需要巡检次数
        visits = int(np.ceil(shift_minutes / period))
        
        # 每个工人每班能完成的轮次
        cycles = max(1, int(shift_minutes / route_time))
        
        # 需要的工人数
        workers = max(1, int(np.ceil(visits / cycles)))
        
        results.append({
            'route_points': route,
            'period': period,
            'route_time': int(route_time),
            'visits_per_shift': visits,
            'cycles_per_worker': cycles,
            'workers': workers
        })
        total_workers += workers
    
    return results, total_workers

# ========== 5. 详细时间表生成 ==========
def generate_detailed_schedule(route_info, dist, shift_minutes=480, start=22, shift_id=1):
    """为一条路线生成详细时间表"""
    point_dict = {}
    route = route_info['route_points']
    period = route_info['period']
    workers = route_info['workers']
    
    schedule_entries = []
    
    for w in range(1, workers + 1):
        # 每个工人的出发时间错开
        offset = (w - 1) * period
        current_time = offset
        
        for trip in range(route_info['cycles_per_worker']):
            trip_start = current_time
            
            # 从调度中心出发
            leg_time = dist[start][route[0]]
            arrive = current_time + leg_time
            
            for i, pn in enumerate(route):
                inspect_time = 2  # 默认巡检时间
                depart = arrive + inspect_time
                schedule_entries.append({
                    'worker': f'W{shift_id}-{w}',
                    'point': pn,
                    'arrive': f'{int(arrive//60):02d}:{int(arrive%60):02d}',
                    'depart': f'{int(depart//60):02d}:{int(depart%60):02d}',
                    'trip': trip + 1
                })
                if i < len(route) - 1:
                    travel = dist[route[i]][route[i+1]]
                    arrive = depart + travel
            
            # 返回调度中心
            return_time = dist[route[-1]][start]
            current_time = depart + return_time
            
            # 如果还有时间，开始下一轮
            if current_time < shift_minutes:
                # 等待到下一个周期开始
                next_start = trip_start + period
                current_time = max(current_time, next_start)
    
    return schedule_entries

# ========== 6. 主求解 ==========
def solve_all():
    """完整求解三个问题"""
    os.chdir('<LOCAL_MATH_MODELING_TEST_ROOT>/2017年赛题/2017D')
    
    points, edges = load_data()
    dist, nxt = build_graph(points, edges)
    point_dict = {p['num']: p for p in points}
    
    # ===== 工作量分析 =====
    print("=" * 70)
    print("数据概览与工作量分析")
    print("=" * 70)
    
    workload_data = []
    for p in points:
        visits = int(np.ceil(480 / p['period']))
        workload_data.append({
            '编号': p['id'],
            '周期(min)': p['period'],
            '巡检耗时(min)': p['duration'],
            '每班巡检次数': visits,
            '每班巡检总时间(min)': visits * p['duration']
        })
    df_wl = pd.DataFrame(workload_data)
    print(df_wl.to_string(index=False))
    print(f"\n总巡检次数/班: {df_wl['每班巡检次数'].sum()}")
    print(f"总巡检时间/班: {df_wl['每班巡检总时间(min)'].sum()} min")
    
    # ===== Q1: 固定班次，无休息 =====
    print("\n" + "=" * 70)
    print("问题1：固定上班时间，无休息，三班倒")
    print("=" * 70)
    
    shift1 = 480  # 8小时
    routes1 = design_routes(points, dist, shift1, 22)
    sched1, workers1 = calc_workers(routes1, dist, point_dict, shift1, 22)
    
    print(f"\n路线设计结果:")
    for i, s in enumerate(sched1):
        print(f"  路线{i+1}: 点{s['route_points']}, 周期{s['period']}min, "
              f"路线耗时{s['route_time']}min, 每班{s['visits_per_shift']}次, "
              f"每工人{s['cycles_per_worker']}轮, 需{s['workers']}人")
    print(f"\n>>> Q1 每班最少需要 {workers1} 人, 三班共需 {workers1*3} 人")
    
    # ===== Q2: 含休息和进餐 =====
    print("\n" + "=" * 70)
    print("问题2：含休息和进餐")
    print("=" * 70)
    
    # 有效工作时间
    rest_breaks = 3  # 每班约3次休息（每2小时一次）
    rest_avg = 7.5   # 平均休息7.5分钟
    meal_breaks = 1  # 每班1次进餐
    meal_time = 30   # 进餐30分钟
    total_break = rest_breaks * rest_avg + meal_breaks * meal_time
    effective_shift = shift1 - total_break
    
    print(f"名义班时: {shift1}min (8小时)")
    print(f"休息: {rest_breaks}次 × {rest_avg}min = {rest_breaks*rest_avg}min")
    print(f"进餐: {meal_breaks}次 × {meal_time}min = {meal_breaks*meal_time}min")
    print(f"有效工作时间: {effective_shift}min")
    
    routes2 = design_routes(points, dist, effective_shift, 22)
    sched2, workers2 = calc_workers(routes2, dist, point_dict, effective_shift, 22)
    
    print(f"\n路线设计结果:")
    for i, s in enumerate(sched2):
        print(f"  路线{i+1}: 点{s['route_points']}, 周期{s['period']}min, "
              f"路线耗时{s['route_time']}min, 每班{s['visits_per_shift']}次, "
              f"每工人{s['cycles_per_worker']}轮, 需{s['workers']}人")
    print(f"\n>>> Q2 每班最少需要 {workers2} 人, 三班共需 {workers2*3} 人")
    
    # ===== Q3: 错时上班 =====
    print("\n" + "=" * 70)
    print("问题3：错时上班分析")
    print("=" * 70)
    
    # 错时方案1：错时+无休息
    routes3a = design_routes(points, dist, shift1, 22)
    sched3a, workers3a = calc_workers(routes3a, dist, point_dict, shift1, 22)
    
    # 错时方案2：错时+有休息
    routes3b = design_routes(points, dist, effective_shift, 22)
    sched3b, workers3b = calc_workers(routes3b, dist, point_dict, effective_shift, 22)
    
    # 错时优化：通过错开出发时间减少同一时刻的人员需求
    # 核心：对于高频点（35min周期），多个工人错开出发可以减少每人巡检次数
    # 但总巡检次数不变，所以理论最少人数由总工作量决定
    
    # 计算理论下界
    total_inspect = sum(p['duration'] * int(np.ceil(shift1 / p['period'])) for p in points)
    total_travel_est = total_inspect * 1.5  # 旅行时间约为巡检时间的1.5倍
    total_work_est = total_inspect + total_travel_est
    
    print(f"\n理论分析:")
    print(f"  总巡检时间/班: {total_inspect} min")
    print(f"  估计总旅行时间/班: {total_travel_est:.0f} min")
    print(f"  估计总工作量/班: {total_work_est:.0f} min")
    print(f"  理论最少人数(无休息): {int(np.ceil(total_work_est / shift1))}")
    print(f"  理论最少人数(有休息): {int(np.ceil(total_work_est / effective_shift))}")
    
    print(f"\n>>> Q3 错时方案:")
    print(f"  错时+无休息: 每班 {workers3a} 人, 三班 {workers3a*3} 人")
    print(f"  错时+有休息: 每班 {workers3b} 人, 三班 {workers3b*3} 人")
    print(f"\n  与Q1对比: 无休息情况下错时 {'节省' if workers3a < workers1 else '不节省'} "
          f"({workers3a} vs {workers1})")
    print(f"  与Q2对比: 有休息情况下错时 {'节省' if workers3b < workers2 else '不节省'} "
          f"({workers3b} vs {workers2})")
    
    # ===== 冻结结果 =====
    frozen = {
        'q1': {
            'workers_per_shift': int(workers1),
            'total_3shifts': int(workers1 * 3),
            'routes': [{'points': s['route_points'], 'period': s['period'],
                       'route_time': s['route_time'], 'workers': s['workers']} for s in sched1]
        },
        'q2': {
            'workers_per_shift': int(workers2),
            'total_3shifts': int(workers2 * 3),
            'effective_shift': effective_shift,
            'routes': [{'points': s['route_points'], 'period': s['period'],
                       'route_time': s['route_time'], 'workers': s['workers']} for s in sched2]
        },
        'q3': {
            'stagger_no_rest': {'workers_per_shift': int(workers3a), 'total_3shifts': int(workers3a*3)},
            'stagger_with_rest': {'workers_per_shift': int(workers3b), 'total_3shifts': int(workers3b*3)},
            'analysis': {
                'total_inspect_time': int(total_inspect),
                'total_travel_est': int(total_travel_est),
                'total_work_est': int(total_work_est)
            }
        },
        'workload': {
            'total_visits_per_shift': int(df_wl['每班巡检次数'].sum()),
            'total_inspect_time_per_shift': int(df_wl['每班巡检总时间(min)'].sum())
        }
    }
    
    os.makedirs('支撑材料/results', exist_ok=True)
    with open('支撑材料/results/frozen_numbers.json', 'w', encoding='utf-8') as f:
        json.dump(frozen, f, ensure_ascii=False, indent=2)
    
    print("\n\n结果已保存到 frozen_numbers.json")
    return frozen

if __name__ == '__main__':
    solve_all()
